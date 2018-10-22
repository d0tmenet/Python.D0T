import pygame
from pygame.locals import *

def main():
    # Initialise screen
    pygame.init()
    screen = pygame.display.set_mode((640, 480))
    pygame.display.set_caption('Basic Pygame program')

    # Fill background
    background = pygame.Surface(screen.get_size())
    background = background.convert()
    background.fill((150, 250, 250))

    # Display some text
    font = pygame.font.Font(None, 36)
    text = font.render("Hello There", 1, (10, 10, 10))
    textpos = text.get_rect()
    textpos.centerx = background.get_rect().centerx
    background.blit(text, textpos)

    # Blit everything to the screen
    screen.blit(background, (0, 0))
    pygame.display.flip()

    # Event loop
    while 1:
        for event in pygame.event.get():
            if event.type == QUIT:
                return

        screen.blit(background, (0, 0))
        pygame.display.flip()


if __name__ == '__main__': main()

from datetime import datetime
now = datetime.now()
mm = str(now.month)
dd = str(now.day)
yyyy = str(now.year)
hour = str(now.hour)
mi = str(now.minute)
ss = str(now.second)
print('Current course in the Universe: ' + mm + "/" + dd + "/" + yyyy + " " + hour + ":" + mi + ":" + ss)

print('Hello Boss')
print('What is your name')
yourName = input()
print('It is good to meet you, ' + yourName)
print('What is your Age, ' + yourName)
yourAge = input()

from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
ws['A2'] = datetime.now()

# Save the file
wb.save("sample.xlsx")